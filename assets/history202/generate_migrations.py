#!/usr/bin/env python3
"""
Generate SQL migration scripts for dental clinic 2023-2025 data.
Steps 6, 7, 8: Patients, Visits/Treatments, Invoices/Payments
"""

import openpyxl
from collections import defaultdict
from datetime import datetime

# ─── Load workbook ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook('Data2023_2025.xlsx')

# ─── Helpers ──────────────────────────────────────────────────────────────────

def sql_str(val):
    """Escape a string value for SQL single-quote embedding."""
    if val is None:
        return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

def clean_phone(val):
    """Convert float phone like 9704770771.0 to string '9704770771'."""
    if val is None:
        return None
    return str(int(float(val)))

def clean_name_part(val):
    """Return None if val is '.', ',', or empty/None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('.', ',', ''):
        return None
    return s

def map_gender(val):
    """Map lowercase gender from Excel to DB enum."""
    if val is None:
        return 'Other'
    v = str(val).strip().lower()
    if v == 'male':
        return 'Male'
    elif v == 'female':
        return 'Female'
    return 'Other'

def map_payment_mode(val):
    """Map payment mode from Excel to DB enum."""
    if val is None:
        return 'Cash'
    v = str(val).strip().lower()
    if v == 'cash':
        return 'Cash'
    elif v in ('online', 'upi'):
        return 'UPI'
    elif v == 'card':
        return 'Card'
    return 'Cash'

def format_date(dt):
    """Format datetime to SQL date string."""
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d')
    return str(dt)[:10]

# ─── Unknown patient code → PRODONT mapping ───────────────────────────────────
UNKNOWN_CODE_MAP = {
    'PT29491852': 'PRODONT338',
    '1288.0':    'PRODONT350',
    'PT29847569': 'PRODONT360',
    'PT30201298': 'PRODONT376',
    'YK192':     'PRODONT392',
    'PT30559517': 'PRODONT395',
    'PT31594809': 'PRODONT418',
}

# ─── Load Patients ─────────────────────────────────────────────────────────────
ws_pat = wb['Patients']
patients = []
email_used = {}  # email -> PRODONT code (for dedup)
phone_email_counter = defaultdict(int)  # phone -> count used

SKIP_PATIENTS = {'PRODONT513'}  # exact dup of PRODONT512

for r in range(2, ws_pat.max_row + 1):
    code = ws_pat.cell(r, 1).value
    if not code:
        continue
    if code in SKIP_PATIENTS:
        continue

    fname = clean_name_part(ws_pat.cell(r, 2).value)
    mname = clean_name_part(ws_pat.cell(r, 3).value)
    lname = clean_name_part(ws_pat.cell(r, 4).value)
    phone_raw = ws_pat.cell(r, 5).value
    email_raw = ws_pat.cell(r, 7).value
    gender_raw = ws_pat.cell(r, 8).value
    dob_raw = ws_pat.cell(r, 13).value

    phone = clean_phone(phone_raw) if phone_raw else None

    # Determine email
    if email_raw and str(email_raw).strip() and '@' in str(email_raw):
        email = str(email_raw).strip().lower()
    elif phone:
        base_email = f'{phone}@prodontics.in'
        # Handle duplicate phone → add suffix for 2nd+ patient with same phone
        phone_email_counter[phone] += 1
        count = phone_email_counter[phone]
        if count == 1:
            email = base_email
        else:
            email = f'{phone}_{count}@prodontics.in'
    else:
        email = f'{code.lower()}@prodontics.in'

    # Check middle name isn't an email address (PRODONT469 issue)
    if mname and '@' in mname:
        mname = None

    patients.append({
        'code': code,
        'fname': fname or '',
        'lname': lname,
        'phone': phone,
        'email': email,
        'gender': map_gender(gender_raw),
        'dob': format_date(dob_raw),
    })

print(f'Loaded {len(patients)} patients (after skips)')

# ─── Load TreatmentPlan ────────────────────────────────────────────────────────
ws_tp = wb['TreatmentPlan']
treatments = []
for r in range(2, ws_tp.max_row + 1):
    proc_id = ws_tp.cell(r, 1).value
    if not proc_id:
        continue
    patient_code_raw = ws_tp.cell(r, 2).value
    treat_date = ws_tp.cell(r, 4).value
    proc_name = ws_tp.cell(r, 5).value
    notes = ws_tp.cell(r, 6).value
    cost = ws_tp.cell(r, 7).value
    discount = ws_tp.cell(r, 8).value
    disc_type = ws_tp.cell(r, 9).value

    if not patient_code_raw or not treat_date:
        continue

    # Resolve patient code
    pcode = str(patient_code_raw)
    if pcode in UNKNOWN_CODE_MAP:
        pcode = UNKNOWN_CODE_MAP[pcode]

    # Skip PRODONT503 rows (dup patient) - should be PRODONT504
    if pcode == 'PRODONT503':
        pcode = 'PRODONT504'

    # Calculate effective cost
    cost_val = float(cost) if cost is not None else 0.0
    disc_val = float(discount) if discount is not None else 0.0
    if disc_type == '%':
        effective_cost = cost_val * (1 - disc_val / 100)
    else:  # INR
        effective_cost = cost_val - disc_val

    treatments.append({
        'patient_code': pcode,
        'visit_date': format_date(treat_date),
        'proc_name': str(proc_name).strip() if proc_name else '',
        'notes': str(notes).strip() if notes else None,
        'cost': cost_val,
        'discount': disc_val,
        'disc_type': disc_type or 'INR',
        'effective_cost': effective_cost,
    })

print(f'Loaded {len(treatments)} treatment rows')

# ─── Build visits (unique patient+date combos) ────────────────────────────────
visits_map = defaultdict(list)
for t in treatments:
    visits_map[(t['patient_code'], t['visit_date'])].append(t)

visits = []
for (pcode, vdate), trows in sorted(visits_map.items()):
    visits.append({
        'patient_code': pcode,
        'visit_date': vdate,
        'treatments': trows,
    })

print(f'Unique visits: {len(visits)}')

# ─── Load Payments ────────────────────────────────────────────────────────────
ws_pay = wb['Payments']
all_payment_rows = []
for r in range(2, ws_pay.max_row + 1):
    pat_name = ws_pay.cell(r, 1).value
    if not pat_name:
        continue
    amount = ws_pay.cell(r, 2).value
    treat_plan = ws_pay.cell(r, 3).value
    invoice_num = ws_pay.cell(r, 5).value
    pay_mode = ws_pay.cell(r, 7).value
    pay_date = ws_pay.cell(r, 8).value

    all_payment_rows.append({
        'patient_name': str(pat_name).strip(),
        'amount': float(amount) if amount else 0.0,
        'treat_plan': str(treat_plan).strip() if treat_plan else '',
        'invoice_num': str(invoice_num).strip() if invoice_num else '',
        'pay_mode': map_payment_mode(pay_mode),
        'pay_date': format_date(pay_date),
    })

print(f'Loaded {len(all_payment_rows)} payment rows')

# ─── Build invoice → patient code mapping ─────────────────────────────────────
# Build patient name → code lookup from patients list
pat_by_code = {p['code']: p for p in patients}

# Name normalization for fuzzy match
def norm_name(n):
    return ' '.join(n.replace('  ', ' ').strip().lower().split())

# Build lookup using all possible name combinations from the raw Excel data
pat_name_to_code = {}
ws_pat2 = wb['Patients']
for r in range(2, ws_pat2.max_row + 1):
    code = ws_pat2.cell(r, 1).value
    if not code or code in SKIP_PATIENTS:
        continue
    fname = ws_pat2.cell(r, 2).value
    mname = ws_pat2.cell(r, 3).value
    lname = ws_pat2.cell(r, 4).value

    fname_s = str(fname).strip() if fname else ''
    mname_s = str(mname).strip() if mname else ''
    lname_s = str(lname).strip() if lname else ''

    # Build all useful name keys
    # 1. fname only
    if fname_s:
        k = norm_name(fname_s)
        pat_name_to_code.setdefault(k, code)

    # 2. fname + lname (if lname not '.' or empty)
    if fname_s and lname_s and lname_s not in ('.', ',', ''):
        k = norm_name(fname_s + ' ' + lname_s)
        pat_name_to_code.setdefault(k, code)

    # 3. fname + mname (if mname not '.' etc.)
    if fname_s and mname_s and mname_s not in ('.', ',', '') and '@' not in mname_s:
        k = norm_name(fname_s + ' ' + mname_s)
        pat_name_to_code.setdefault(k, code)

    # 4. fname + mname + lname (all three)
    if fname_s and mname_s and lname_s and mname_s not in ('.', ',', '') and lname_s not in ('.', ',', '') and '@' not in mname_s:
        k = norm_name(fname_s + ' ' + mname_s + ' ' + lname_s)
        pat_name_to_code.setdefault(k, code)

# Build invoices grouped by invoice number
invoices_raw = defaultdict(list)
for row in all_payment_rows:
    invoices_raw[row['invoice_num']].append(row)

# For each invoice, find the patient code
# INVPRODONTXXX: patient is already in this sheet (confirmed above), find by name
# INV10X: find by name too

SKIP_INVOICES = {'INV101'}  # Ramya G cannot be matched

# Build invoice → patient_code lookup
invoice_patient_map = {}  # invoice_num → patient_code

for inv, rows in invoices_raw.items():
    if inv in SKIP_INVOICES:
        continue

    pname_raw = rows[0]['patient_name']
    pname_norm = norm_name(pname_raw)

    # Direct lookup
    code = pat_name_to_code.get(pname_norm)

    # Fuzzy fallback 1: strip trailing ' .' from payment name (patient has dot as last name)
    if not code and pname_norm.endswith(' .'):
        code = pat_name_to_code.get(pname_norm[:-2].strip())

    # Fuzzy fallback 2: try without middle word (first + last)
    if not code:
        parts = pname_norm.split()
        if len(parts) >= 3:
            short = parts[0] + ' ' + parts[-1]
            code = pat_name_to_code.get(short)

    # Fuzzy fallback 3: try first word only
    if not code:
        parts = pname_norm.split()
        if parts:
            code = pat_name_to_code.get(parts[0])

    # Fuzzy fallback 4: for name after stripping trailing dot, try first word only
    if not code and pname_norm.endswith(' .'):
        stripped = pname_norm[:-2].strip()
        parts = stripped.split()
        if parts:
            code = pat_name_to_code.get(parts[0])

    if not code:
        # Special cases: INVPRODONT197 → PRODONT503 in name but we skip 503 → use 504
        if 'baddepudi venkata suresh kumar' in pname_norm:
            code = 'PRODONT504'

    if not code:
        print(f'  WARNING: Cannot match invoice {inv} patient "{pname_raw}" to any code')
        continue

    # Remap skipped patients: PRODONT503 → PRODONT504
    if code == 'PRODONT503':
        code = 'PRODONT504'

    invoice_patient_map[inv] = code

print(f'Matched {len(invoice_patient_map)} invoices to patient codes')

# ─── Build invoice details ─────────────────────────────────────────────────────
# For each invoice, we need: invoice_num, patient_code, total, date, pay_mode
# And find the visit (patient+date) to link to

# Build visit date lookup: patient_code → list of visit dates
visit_dates_by_patient = defaultdict(list)
for v in visits:
    visit_dates_by_patient[v['patient_code']].append(v['visit_date'])

def find_nearest_visit(patient_code, pay_date):
    """Find the visit date for this patient nearest to pay_date."""
    dates = visit_dates_by_patient.get(patient_code, [])
    if not dates:
        return None
    if pay_date in dates:
        return pay_date
    # Find nearest
    from datetime import date as dclass
    try:
        pd = datetime.strptime(pay_date, '%Y-%m-%d').date()
        nearest = min(dates, key=lambda d: abs((datetime.strptime(d, '%Y-%m-%d').date() - pd).days))
        return nearest
    except Exception:
        return dates[0]

invoice_details = []
for inv, rows in invoices_raw.items():
    if inv in SKIP_INVOICES or inv not in invoice_patient_map:
        continue

    pcode = invoice_patient_map[inv]
    pay_date = rows[0]['pay_date']
    total = sum(r['amount'] for r in rows)
    # Deduplicate payment rows (same amount+treatment = identical rows)
    # Only keep unique (amount, treat_plan, pay_date) combos
    seen_pay = set()
    unique_rows = []
    for r in rows:
        key = (r['amount'], r['treat_plan'], r['pay_date'], r['pay_mode'])
        if key not in seen_pay:
            seen_pay.add(key)
            unique_rows.append(r)

    # Pay mode: use first row's mode
    pay_mode = unique_rows[0]['pay_mode']

    # Find visit to link invoice to
    nearest_visit = find_nearest_visit(pcode, pay_date)

    invoice_details.append({
        'invoice_num': inv,
        'patient_code': pcode,
        'total': total,
        'pay_date': pay_date,
        'pay_mode': pay_mode,
        'visit_date': nearest_visit,
        'payment_rows': unique_rows,
    })

print(f'Built {len(invoice_details)} invoice details')

# ─── Verify unique payment rows total ────────────────────────────────────────
total_pay_rows = sum(len(inv['payment_rows']) for inv in invoice_details)
print(f'Total unique payment rows across all invoices: {total_pay_rows}')


# ═══════════════════════════════════════════════════════════════════════════════
# GENERATE FILE 1: migration_step6_patients_2023_2025.sql
# ═══════════════════════════════════════════════════════════════════════════════

step6_lines = []
step6_lines.append("""-- =============================================================================
-- migration_step6_patients_2023_2025.sql
-- =============================================================================
-- Inserts patients PRODONT311 through PRODONT521 (skipping PRODONT503, PRODONT513)
-- into auth.users, auth.identities, and public.patients.
--
-- Patient count : {count} patients
-- Run order     : Step 6 (after Steps 1-5)
-- Safe to re-run: YES (all inserts are guarded with WHERE NOT EXISTS / ON CONFLICT)
-- =============================================================================

DO $$
DECLARE
  v_user_id   UUID;
  v_email     TEXT;
  v_phone     TEXT;
  v_code      TEXT;
  v_fname     TEXT;
  v_lname     TEXT;
  v_gender    TEXT;
  v_dob       DATE;

  -- Each record: (legacy_code, first_name, last_name, phone, email, gender, dob_or_null)
  TYPE patient_rec IS RECORD (
    code   TEXT,
    fname  TEXT,
    lname  TEXT,
    phone  TEXT,
    email  TEXT,
    gender TEXT,
    dob    TEXT
  );
  rec patient_rec;

  patients patient_rec[] := ARRAY[""".format(count=len(patients)))

# Build the patient array
pat_rows = []
for p in patients:
    fname_sql = sql_str(p['fname'])
    lname_sql = sql_str(p['lname'])
    phone_sql = sql_str(p['phone'])
    email_sql = sql_str(p['email'])
    gender_sql = sql_str(p['gender'])
    dob_sql = sql_str(p['dob'])
    code_sql = sql_str(p['code'])
    pat_rows.append(f"    ROW({code_sql}, {fname_sql}, {lname_sql}, {phone_sql}, {email_sql}, {gender_sql}, {dob_sql})::patient_rec")

step6_lines.append(',\n'.join(pat_rows))
step6_lines.append("""  ];

BEGIN
  FOREACH rec IN ARRAY patients LOOP
    v_code   := rec.code;
    v_fname  := rec.fname;
    v_lname  := rec.lname;
    v_phone  := rec.phone;
    v_email  := rec.email;
    v_gender := rec.gender;
    v_dob    := CASE WHEN rec.dob IS NOT NULL THEN rec.dob::DATE ELSE NULL END;

    -- Check if patient already exists (idempotent)
    IF EXISTS (SELECT 1 FROM public.patients WHERE legacy_patient_code = v_code) THEN
      CONTINUE;
    END IF;

    -- Generate a new UUID for this patient
    v_user_id := gen_random_uuid();

    -- Insert into auth.users (if email not already registered)
    IF NOT EXISTS (SELECT 1 FROM auth.users WHERE email = v_email) THEN
      INSERT INTO auth.users (
        id, email,
        encrypted_password,
        raw_app_meta_data,
        raw_user_meta_data,
        aud, role,
        created_at, updated_at
      ) VALUES (
        v_user_id,
        v_email,
        '',
        '{"provider":"email","providers":["email"]}',
        '{}',
        'authenticated',
        'authenticated',
        now(), now()
      );
    ELSE
      -- Reuse existing auth user id
      SELECT id INTO v_user_id FROM auth.users WHERE email = v_email LIMIT 1;
    END IF;

    -- Insert into auth.identities
    IF NOT EXISTS (SELECT 1 FROM auth.identities WHERE provider_id = v_email AND provider = 'email') THEN
      INSERT INTO auth.identities (
        id, user_id, provider, provider_id,
        identity_data,
        created_at, updated_at
      ) VALUES (
        gen_random_uuid(),
        v_user_id,
        'email',
        v_email,
        jsonb_build_object('sub', v_user_id::text, 'email', v_email),
        now(), now()
      );
    END IF;

    -- Insert into public.patients
    IF NOT EXISTS (SELECT 1 FROM public.patients WHERE legacy_patient_code = v_code) THEN
      INSERT INTO public.patients (
        id, auth_user_id,
        legacy_patient_code,
        first_name, last_name,
        phone, email,
        gender,
        date_of_birth,
        created_at
      ) VALUES (
        v_user_id, v_user_id,
        v_code,
        v_fname, v_lname,
        v_phone, v_email,
        v_gender::gender_type,
        v_dob,
        now()
      );
    END IF;

  END LOOP;
END $$;

-- Verification
SELECT
  COUNT(*)                                          AS total_patients,
  COUNT(*) FILTER (WHERE legacy_patient_code LIKE 'PRODONT3%') AS prodont3xx,
  COUNT(*) FILTER (WHERE legacy_patient_code LIKE 'PRODONT4%') AS prodont4xx,
  COUNT(*) FILTER (WHERE legacy_patient_code LIKE 'PRODONT5%') AS prodont5xx
FROM public.patients
WHERE legacy_patient_code >= 'PRODONT311'
  AND legacy_patient_code <= 'PRODONT521';
""")

step6_sql = ''.join(step6_lines)

# Write file
with open('migration_step6_patients_2023_2025.sql', 'w', encoding='utf-8') as f:
    f.write(step6_sql)

print(f'\nWrote migration_step6_patients_2023_2025.sql ({len(step6_sql)} bytes)')


# ═══════════════════════════════════════════════════════════════════════════════
# GENERATE FILE 2: migration_step7_visits_2023_2025.sql
# ═══════════════════════════════════════════════════════════════════════════════

step7_lines = []
step7_lines.append(f"""-- =============================================================================
-- migration_step7_visits_2023_2025.sql
-- =============================================================================
-- Inserts visits and treatment_plans for 2023-2025 data.
--
-- Visits count         : {len(visits)}
-- Treatment rows count : {len(treatments)}
-- Run order            : Step 7 (after Step 6)
-- Safe to re-run       : YES (WHERE NOT EXISTS guards)
-- =============================================================================

DO $$
DECLARE
  v_doctor_id  UUID;
  v_patient_id UUID;
  v_visit_id   UUID;
BEGIN
  -- Resolve doctor once
  SELECT id INTO v_doctor_id
  FROM doctors
  WHERE first_name ILIKE '%Surya%'
  LIMIT 1;

  IF v_doctor_id IS NULL THEN
    RAISE EXCEPTION 'Doctor with name Surya not found in doctors table';
  END IF;

""")

# Emit each visit as a block
for v in visits:
    pcode = v['patient_code']
    vdate = v['visit_date']
    trows = v['treatments']

    step7_lines.append(f"""  -- Visit: {pcode} on {vdate}
  SELECT id INTO v_patient_id FROM public.patients WHERE legacy_patient_code = {sql_str(pcode)} LIMIT 1;
  IF v_patient_id IS NULL THEN
    RAISE WARNING 'Patient % not found, skipping visit on {vdate}', {sql_str(pcode)};
  ELSE
    -- Insert visit if not exists
    SELECT id INTO v_visit_id
    FROM public.visits
    WHERE patient_id = v_patient_id
      AND visit_date::date = {sql_str(vdate)}::date
    LIMIT 1;

    IF v_visit_id IS NULL THEN
      v_visit_id := gen_random_uuid();
      INSERT INTO public.visits (
        id, patient_id, doctor_id,
        visit_date, status,
        created_at, updated_at
      ) VALUES (
        v_visit_id,
        v_patient_id,
        v_doctor_id,
        {sql_str(vdate)}::timestamptz,
        'complete',
        now(), now()
      );
    END IF;

""")

    # Emit treatment plans for this visit
    for t in trows:
        notes_sql = sql_str(t['notes'])
        cost_int = int(round(t['effective_cost']))
        step7_lines.append(f"""    -- Treatment: {t['proc_name'][:60]}
    IF NOT EXISTS (
      SELECT 1 FROM public.treatment_plans
      WHERE visit_id = v_visit_id
        AND treatment_name = {sql_str(t['proc_name'])}
        AND COALESCE(notes, '') = COALESCE({notes_sql}, '')
    ) THEN
      INSERT INTO public.treatment_plans (
        id, visit_id,
        treatment_name, notes,
        cost,
        status,
        created_at, updated_at
      ) VALUES (
        gen_random_uuid(),
        v_visit_id,
        {sql_str(t['proc_name'])},
        {notes_sql},
        {cost_int},
        'planned',
        now(), now()
      );
    END IF;

""")

    step7_lines.append("  END IF;\n\n")

step7_lines.append("""END $$;

-- Verification
SELECT
  (SELECT COUNT(*) FROM public.visits
   WHERE created_at >= NOW() - INTERVAL '1 day') AS visits_inserted_today,
  (SELECT COUNT(*) FROM public.treatment_plans
   WHERE created_at >= NOW() - INTERVAL '1 day') AS treatments_inserted_today;
""")

step7_sql = ''.join(step7_lines)

with open('migration_step7_visits_2023_2025.sql', 'w', encoding='utf-8') as f:
    f.write(step7_sql)

print(f'Wrote migration_step7_visits_2023_2025.sql ({len(step7_sql)} bytes)')


# ═══════════════════════════════════════════════════════════════════════════════
# GENERATE FILE 3: migration_step8_invoices_payments_2023_2025.sql
# ═══════════════════════════════════════════════════════════════════════════════

step8_lines = []
step8_lines.append(f"""-- =============================================================================
-- migration_step8_invoices_payments_2023_2025.sql
-- =============================================================================
-- Inserts invoices and payments for 2023-2025 data.
--
-- Invoices count        : {len(invoice_details)}  (INV101 skipped - Ramya G unmatchable)
-- Total payment rows    : {total_pay_rows}
-- Run order             : Step 8 (after Steps 6 and 7)
-- Safe to re-run        : YES (ON CONFLICT / WHERE NOT EXISTS guards)
-- =============================================================================

DO $$
DECLARE
  v_patient_id UUID;
  v_visit_id   UUID;
  v_invoice_id UUID;
BEGIN

""")

for inv in invoice_details:
    inv_num = inv['invoice_num']
    pcode = inv['patient_code']
    total = int(round(inv['total']))
    pay_date = inv['pay_date']
    pay_mode = inv['pay_mode']
    visit_date = inv['visit_date']

    step8_lines.append(f"""  -- Invoice: {inv_num} | Patient: {pcode} | Date: {pay_date} | Total: {total}
  SELECT id INTO v_patient_id FROM public.patients WHERE legacy_patient_code = {sql_str(pcode)} LIMIT 1;

  IF v_patient_id IS NULL THEN
    RAISE WARNING 'Patient % not found, skipping invoice {inv_num}', {sql_str(pcode)};
  ELSE
    -- Find the linked visit
    SELECT id INTO v_visit_id
    FROM public.visits
    WHERE patient_id = v_patient_id
      AND visit_date::date = {sql_str(visit_date)}::date
    LIMIT 1;

    -- Insert invoice
    INSERT INTO public.invoices (
      id, visit_id, patient_id,
      legacy_bill_number,
      total_amount, amount_paid, balance,
      payment_mode,
      invoice_date,
      created_at, updated_at
    )
    SELECT
      gen_random_uuid(),
      v_visit_id,
      v_patient_id,
      {sql_str(inv_num)},
      {total},
      {total},
      0,
      {sql_str(pay_mode)}::payment_mode_type,
      {sql_str(pay_date)}::date,
      now(), now()
    WHERE NOT EXISTS (
      SELECT 1 FROM public.invoices
      WHERE legacy_bill_number = {sql_str(inv_num)}
    );

    SELECT id INTO v_invoice_id FROM public.invoices WHERE legacy_bill_number = {sql_str(inv_num)} LIMIT 1;

""")

    # Payment rows for this invoice
    for pr in inv['payment_rows']:
        note_text = pr['treat_plan'][:200] if pr['treat_plan'] else ''
        step8_lines.append(f"""    -- Payment: {pr['pay_mode']} {int(round(pr['amount']))} for {note_text[:50]}
    IF v_invoice_id IS NOT NULL AND v_visit_id IS NOT NULL THEN
      IF NOT EXISTS (
        SELECT 1 FROM public.payments
        WHERE visit_id = v_visit_id
          AND notes = {sql_str(note_text)}
          AND amount = {int(round(pr['amount']))}
          AND payment_date::date = {sql_str(pr['pay_date'])}::date
      ) THEN
        INSERT INTO public.payments (
          id, invoice_id, visit_id, patient_id,
          amount, payment_mode,
          notes,
          payment_date,
          created_at, updated_at
        ) VALUES (
          gen_random_uuid(),
          v_invoice_id,
          v_visit_id,
          v_patient_id,
          {int(round(pr['amount']))},
          {sql_str(pr['pay_mode'])}::payment_mode_type,
          {sql_str(note_text)},
          {sql_str(pr['pay_date'])}::date,
          now(), now()
        );
      END IF;
    END IF;

""")

    step8_lines.append("  END IF;\n\n")

step8_lines.append("""END $$;

-- Verification
SELECT
  (SELECT COUNT(*) FROM public.invoices
   WHERE created_at >= NOW() - INTERVAL '1 day') AS invoices_inserted_today,
  (SELECT COUNT(*) FROM public.payments
   WHERE created_at >= NOW() - INTERVAL '1 day') AS payments_inserted_today;
""")

step8_sql = ''.join(step8_lines)

with open('migration_step8_invoices_payments_2023_2025.sql', 'w', encoding='utf-8') as f:
    f.write(step8_sql)

print(f'Wrote migration_step8_invoices_payments_2023_2025.sql ({len(step8_sql)} bytes)')
print('\nAll 3 migration files generated successfully.')
